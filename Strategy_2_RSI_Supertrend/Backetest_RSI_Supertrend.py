import numpy as np
import pandas as pd
import mysql.connector as mdb
import os
from urllib.parse import quote_plus
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Database Configuration
db_config = {
    "host": "localhost",
    "user": "Algo_Trading",
    "password": "Apple@1234",
    "database": "Historical_data_2024"
}

# Connect to MySQL
def connect_db():
    try:
        con = mdb.connect(**db_config)
        print("✅ Database connection successful!")
        return con
    except mdb.Error as e:
        print(f"❌ Database connection error: {e}")
        return None

con = connect_db()
ENCODED_PASS = quote_plus(db_config["password"])
ENGINE = create_engine(f"mysql+mysqlconnector://{db_config['user']}:{ENCODED_PASS}@{db_config['host']}/{db_config['database']}")

# Load Historical Data
def load_data_from_db():
    query = """
        SELECT sym.ticker, dp.price_date AS timestamp, dp.open_price AS open, dp.high_price AS high, 
               dp.low_price AS low, dp.close_price AS close, dp.volume AS volume
        FROM daily_price AS dp
        INNER JOIN symbol AS sym ON dp.symbol_id = sym.id
        ORDER BY sym.ticker, dp.price_date ASC
    """
    df = pd.read_sql(query, con=ENGINE)
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    return df.set_index(["ticker", "timestamp"])

# Calculate RSI
def calculate_rsi(data, period=14):
    delta = data["close"].groupby(level=0).diff()
    gain = delta.where(delta > 0, 0).groupby(level=0).rolling(window=period).mean().reset_index(0, drop=True)
    loss = -delta.where(delta < 0, 0).groupby(level=0).rolling(window=period).mean().reset_index(0, drop=True)
    rs = gain / loss
    data["RSI"] = 100 - (100 / (1 + rs))
    return data

# Calculate Supertrend
def calculate_supertrend(data, period=10, multiplier=3):
    data["hl2"] = (data["high"] + data["low"]) / 2
    data["atr"] = data["high"].rolling(period).max() - data["low"].rolling(period).min()
    data["upper_band"] = data["hl2"] + (multiplier * data["atr"])
    data["lower_band"] = data["hl2"] - (multiplier * data["atr"])
    data["Supertrend"] = np.where(data["close"] > data["upper_band"], data["lower_band"], data["upper_band"])
    return data

# Generate Trading Signals
def generate_signals(data):
    data["Signal"] = 0
    for ticker, group in data.groupby(level=0):
        buy_signals = (group["RSI"] > 60) & (group["close"] > group["Supertrend"])
        data.loc[buy_signals.index, "Signal"] = 1
    return data

# Backtest Strategy (Fixes Applied)
def backtest(data, initial_balance=10000, brokerage_fee=20):
    all_trades = []
    
    serial_no = 1
    total_trades = 0
    profitable_trades = 0
    loss_trades = 0

    balance = initial_balance  
    positions = {}  # Track open positions per stock

    for index, row in data.iterrows():
        ticker = index[0]
        timestamp = index[1]  # Extract timestamp

        if ticker not in positions:
            positions[ticker] = {"shares": 0, "entry_price": 0}

        # Sell Condition
        if positions[ticker]["shares"] > 0 and row["close"] < row["Supertrend"]:
            exit_price = row["close"]
            sell_quantity = positions[ticker]["shares"]
            purchase_value = positions[ticker]["entry_price"] * sell_quantity
            gross_profit = (exit_price - positions[ticker]["entry_price"]) * sell_quantity
            gross_profit_pct = (gross_profit / purchase_value) * 100
            fees = brokerage_fee + (sell_quantity * 0.0005 * exit_price)
            net_profit = gross_profit - fees
            net_profit_pct = (net_profit / initial_balance) * 100

            if net_profit > 0:
                profitable_trades += 1
            else:
                loss_trades += 1

            balance += sell_quantity * exit_price - fees  # Correct balance update

            all_trades.append([
                serial_no, ticker, timestamp, purchase_value, positions[ticker]["entry_price"], sell_quantity,
                exit_price, sell_quantity, gross_profit, gross_profit_pct, fees, net_profit, balance, net_profit_pct
            ])
            serial_no += 1
            total_trades += 1

            positions[ticker] = {"shares": 0, "entry_price": 0}

        # Buy Condition (Buy max shares with available balance)
        if positions[ticker]["shares"] == 0 and row["Signal"] == 1:
            buy_quantity = int((balance - brokerage_fee) / row["close"])  # Ensure brokerage is covered
            if buy_quantity > 0:
                positions[ticker]["shares"] = buy_quantity
                positions[ticker]["entry_price"] = row["close"]
                balance -= buy_quantity * row["close"] + brokerage_fee  # Deduct purchase + brokerage

    # Summary
    summary = [
        ["Total Trades", total_trades],
        ["Profitable Trades", profitable_trades],
        ["Loss Trades", loss_trades],
        ["Final Balance", balance]
    ]

    return all_trades, summary

# Save Results to Excel
def save_results_to_files(all_trades, summary):
    output_dir = "/Users/ankursaraswat/Fyers_API_trading_bot/Strategy_2_RSI_Supertrend/"
    os.makedirs(output_dir, exist_ok=True)

    results_df = pd.DataFrame(all_trades, columns=[
        "Ser No.", "Stock", "Timestamp", "Purchase Value", "Buy Price", "Buy Quantity",
        "Sell Price", "Sell Quantity", "Gross Profit", "Gross Profit %",
        "Fees", "Net P/L", "Fund Available", "Net P/L %"
    ])

    summary_df = pd.DataFrame(summary, columns=["Metric", "Value"])

    excel_path = os.path.join(output_dir, "backtest_results.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="Trades", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Apply Color Coding
    wb = load_workbook(excel_path)
    ws = wb["Trades"]
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):  # Net P/L column
        for cell in row:
            if cell.value < 0:
                cell.fill = red_fill
            elif cell.value > 0:
                cell.fill = green_fill

    wb.save(excel_path)
    print(f"✅ Backtest results saved at: {excel_path}")

# Run Backtest
if __name__ == "__main__":
    df = load_data_from_db()
    df = calculate_rsi(df)
    df = calculate_supertrend(df)
    df = generate_signals(df)
    
    all_trades, summary = backtest(df)
    save_results_to_files(all_trades, summary)
